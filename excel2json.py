import openpyxl
import json

book = openpyxl.load_workbook('a.xlsx')
sheets = book['s2']
path_w = 'output.json'
fruits = []
for i in range(2,126):
	a1 = sheets.cell(row=i, column=1).value
	a2 = sheets.cell(row=i, column=2).value

	a3 = sheets.cell(row=i, column=3).value



	a4 = sheets.cell(row=i, column=4).value


	a5 = sheets.cell(row=i, column=5).value
	a6 = sheets.cell(row=i, column=6).value
	a7 = sheets.cell(row=i, column=7).value
	a8 = sheets.cell(row=i, column=8).value
	a9 = sheets.cell(row=i, column=9).value


	fruits.append({
	                    "nom": a1,
	                    "text": a2,
	                    "A": a3 ,
                        "B": a4 ,
                        "C": a5 ,
                        "D": a6 ,
                        "E": a7 ,
                        "target": a8 ,
                        "correct": a9 ,
             }
    )
            
	
	with open(path_w, mode = 'w', encoding = 'utf-8') as f:
		f.write(json.dumps(fruits, sort_keys=False, ensure_ascii=False,indent=4))